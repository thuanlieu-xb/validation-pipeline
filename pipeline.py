"""
Dagster Pipeline - RAG Training + PO Validation (Google Gemini)
================================================================
Jobs:
  - index_documents_job  : Doc file data/, tao embeddings local, luu ChromaDB
  - po_validation_job    : Doc PO Excel, validate tung PO voi Gemini, luu ket qua

Chay:
  dagster dev -f pipeline.py
"""

import os
import glob
from typing import List, Dict
from datetime import datetime

import dagster as dg
import pandas as pd
from docx import Document
from dotenv import load_dotenv
from google import genai as google_genai
from groq import Groq
import chromadb

load_dotenv()

# ── Paths ─────────────────────────────────────────────────
DATA_DIR       = os.path.join(os.path.dirname(__file__), "data")
CHROMA_DIR     = os.path.join(os.path.dirname(__file__), "chroma_db")
VALIDATION_DIR = os.path.join(os.path.dirname(__file__), "validation_data", "purchase_order")

COLLECTION_NAME = "company_docs"
CHUNK_SIZE      = 500
CHUNK_OVERLAP   = 50

PO_HEADER_COLS = [
    "Buyer/Department", "Buyer", "Display Name", "Business Type",
    "Vendor", "Procurement Type", "Billing Journal",
    "Subcontractor Contract Project",
    "Deliver To", "Taxes",
]


# ══════════════════════════════════════════════════════════
# JOB 1: INDEX DOCUMENTS (RAG Training)
# ══════════════════════════════════════════════════════════

@dg.asset(group_name="training")
def load_documents(context: dg.AssetExecutionContext) -> List[Dict]:
    """Doc tat ca .docx va .xlsx trong data/ va trich xuat text."""
    documents = []

    for path in glob.glob(os.path.join(DATA_DIR, "*.docx")):
        filename = os.path.basename(path)
        context.log.info(f"Doc Word: {filename}")
        try:
            doc = Document(path)
            content = "\n".join(p.text.strip() for p in doc.paragraphs if p.text.strip())
            documents.append({"source": filename, "type": "docx", "content": content})
            context.log.info(f"  -> {len(content)} ky tu")
        except Exception as e:
            context.log.warning(f"  -> Bo qua (file loi hoac rong): {e}")

    for path in glob.glob(os.path.join(DATA_DIR, "*.md")):
        filename = os.path.basename(path)
        context.log.info(f"Doc Markdown: {filename}")
        try:
            with open(path, encoding="utf-8") as f:
                content = f.read().strip()
            documents.append({"source": filename, "type": "md", "content": content})
            context.log.info(f"  -> {len(content)} ky tu")
        except Exception as e:
            context.log.warning(f"  -> Bo qua (file loi): {e}")

    for path in glob.glob(os.path.join(DATA_DIR, "*.xlsx")):
        filename = os.path.basename(path)
        context.log.info(f"Doc Excel: {filename}")
        xl = pd.ExcelFile(path)
        all_text = []
        for sheet in xl.sheet_names:
            df = xl.parse(sheet).fillna("")
            all_text.append(f"[Sheet: {sheet}]")
            for _, row in df.iterrows():
                row_text = " | ".join(str(v) for v in row.values if str(v).strip())
                if row_text:
                    all_text.append(row_text)
        content = "\n".join(all_text)
        documents.append({"source": filename, "type": "xlsx", "content": content})
        context.log.info(f"  -> {len(all_text)} dong")

    context.log.info(f"Tong: {len(documents)} file")
    return documents


@dg.asset(group_name="training")
def chunk_documents(context: dg.AssetExecutionContext, load_documents: List[Dict]) -> List[Dict]:
    """Chia tai lieu thanh cac chunks nho."""
    chunks = []
    for doc in load_documents:
        text, source, start, idx = doc["content"], doc["source"], 0, 0
        while start < len(text):
            chunk_text = text[start : start + CHUNK_SIZE].strip()
            if chunk_text:
                chunks.append({"id": f"{source}_chunk_{idx}", "source": source, "chunk_index": idx, "content": chunk_text})
                idx += 1
            start += CHUNK_SIZE - CHUNK_OVERLAP

    context.log.info(f"Tong: {len(chunks)} chunks")
    return chunks


@dg.asset(group_name="training")
def index_to_chromadb(context: dg.AssetExecutionContext, chunk_documents: List[Dict]) -> str:
    """Tao embeddings bang Gemini API (text-embedding-004) va luu vao ChromaDB."""
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key or api_key.startswith("your-"):
        raise ValueError("Chua cau hinh GEMINI_API_KEY trong file .env")

    client = google_genai.Client(api_key=api_key)
    chroma = chromadb.PersistentClient(path=CHROMA_DIR)

    try:
        chroma.delete_collection(COLLECTION_NAME)
        context.log.info("Da xoa collection cu.")
    except Exception:
        pass

    collection = chroma.create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine"},
    )

    BATCH_SIZE = 20  # Gemini embedding free tier ~100 req/min
    total = len(chunk_documents)

    for i in range(0, total, BATCH_SIZE):
        batch = chunk_documents[i : i + BATCH_SIZE]
        context.log.info(f"Tao Gemini embeddings batch {i//BATCH_SIZE + 1}/{(total - 1)//BATCH_SIZE + 1} ({len(batch)} chunks)...")

        embeddings = []
        for chunk in batch:
            result = client.models.embed_content(
                model="models/gemini-embedding-001",
                contents=chunk["content"],
            )
            embeddings.append(result.embeddings[0].values)

        collection.add(
            ids=[c["id"] for c in batch],
            embeddings=embeddings,
            documents=[c["content"] for c in batch],
            metadatas=[{"source": c["source"], "chunk_index": c["chunk_index"]} for c in batch],
        )

    count = collection.count()
    context.log.info(f"Hoan tat! {count} chunks voi Gemini embeddings trong ChromaDB.")
    return f"Indexed {count} chunks (model: models/gemini-embedding-001)"


# ══════════════════════════════════════════════════════════
# JOB 2: PO VALIDATION (Google Gemini - Mien phi)
# ══════════════════════════════════════════════════════════

@dg.asset(group_name="po_validation")
def load_po_data(context: dg.AssetExecutionContext) -> List[Dict]:
    """Loop tat ca file Excel trong validation_data/, forward-fill, tra ve danh sach records."""
    all_records = []

    xlsx_files = [
        f for f in glob.glob(os.path.join(VALIDATION_DIR, "*.xlsx"))
        if "result" not in os.path.basename(f).lower()  # bo qua file ket qua cu
    ]

    if not xlsx_files:
        raise FileNotFoundError(f"Khong tim thay file Excel nao trong {VALIDATION_DIR}")

    context.log.info(f"Tim thay {len(xlsx_files)} file Excel:")
    for path in xlsx_files:
        filename = os.path.basename(path)
        context.log.info(f"  Doc: {filename}")

        df = pd.read_excel(path, sheet_name=0)
        context.log.info(f"    -> {len(df)} dong goc")

        ffill_cols = [c for c in PO_HEADER_COLS if c in df.columns]
        df[ffill_cols] = df[ffill_cols].ffill()

        df["_source_file"] = filename
        records = df.where(pd.notna(df), None).to_dict(orient="records")
        all_records.extend(records)

        pos = df["Display Name"].dropna().unique() if "Display Name" in df.columns else []
        context.log.info(f"    -> {len(pos)} PO: {list(pos)}")

    context.log.info(f"Tong cong: {len(all_records)} order lines tu {len(xlsx_files)} file")
    return all_records


@dg.asset(group_name="po_validation")
def build_scenarios(context: dg.AssetExecutionContext, load_po_data: List[Dict]) -> List[Dict]:
    """Gom order lines theo PO, tao 1 prompt validation cho ca PO."""
    po_groups: Dict[str, List[Dict]] = {}
    for row in load_po_data:
        po_num = row.get("Display Name") or "UNKNOWN"
        po_groups.setdefault(po_num, []).append(row)

    context.log.info(f"Tim thay {len(po_groups)} PO")
    scenarios = []

    for po_num, lines in po_groups.items():
        header = lines[0]

        # Gom tat ca analytic accounts cua PO nay
        all_analytics = [
            str(line.get('Order Lines/Distribution Analytic Account/Analytic Account', ''))
            for line in lines
            if line.get('Order Lines/Distribution Analytic Account/Analytic Account')
        ]
        analytic_summary = " | ".join(dict.fromkeys(all_analytics)) or "N/A"

        order_lines_text = ""
        for i, line in enumerate(lines, 1):
            order_lines_text += (
                f"\n  Dòng {i}:"
                f"\n    - Sản phẩm                                        : {line.get('Order Lines/Product', 'N/A')}"
                f"\n    - Đơn giá                                         : {line.get('Order Lines/Unit Price', 'N/A')}"
                f"\n    - Thuế dòng                                       : {line.get('Order Lines/Taxes', 'N/A')}"
                f"\n    - Order Lines/Distribution Analytic Account/Analytic Account: {line.get('Order Lines/Distribution Analytic Account/Analytic Account', 'N/A')}"
            )

        prompt = f"""Hãy kiểm tra Purchase Order dưới đây theo đúng quy trình 3 bước.

=== DỮ LIỆU PO ===
- Mã PO                    : {header.get('Display Name', 'N/A')}
- Phòng ban                : {header.get('Buyer/Department', 'N/A')}
- Người mua                : {header.get('Buyer', 'N/A')}
- Vendor                   : {header.get('Vendor', 'N/A')}
- Business Type            : {header.get('Business Type', 'N/A')}
- Procurement Type         : {header.get('Procurement Type', 'N/A')}
- Billing Journal          : {header.get('Billing Journal', 'N/A')}
- Subcontractor Contract   : {header.get('Subcontractor Contract Project', 'N/A')}
- Deliver To               : {header.get('Deliver To', 'N/A')}
- Taxes (PO header)        : {header.get('Taxes', 'N/A')}
- Order Lines/Distribution Analytic Account/Analytic Account: {analytic_summary}

=== ORDER LINES ({len(lines)} dòng) ==={order_lines_text}

=== YÊU CẦU KIỂM TRA ===

BƯỚC 1 — TÍNH ĐIỂM CONFIDENCE (xác định nghiệp vụ thực tế):
Chấm điểm dựa trên các tín hiệu sau:
  +3đ: Subcontractor Contract có dữ liệu → gần chắc chắn XL/ĐT
  +3đ: Analytic Distribution chứa mã BU (XL*, DT*, KD*...)
  +3đ: Tax name chứa keyword đúng nhóm (GTGT Xây Lắp / GTGT TSCĐ / GTGT KDVT...)
  +2đ: Deliver To chứa "NVL", "CCDC", hoặc "Hàng hóa"
  +2đ: Billing Journal thuộc nhóm đúng theo bảng ma trận
  +1đ: Business Type = Construction / Trading
→ Tổng điểm và kết luận: ≥9 = Rất chắc chắn | 6–8 = Khả năng cao | ≤5 = Cần xác nhận

BƯỚC 2 — KIỂM TRA 4 RÀNG BUỘC:

[1] Billing Journal × BU:
  - Journal có khớp với nghiệp vụ xác định từ bước 1 không?
  - Lỗi cross-BU: Journal KDVT nhưng Analytic là XL* → Critical
  - "Nhật ký vay" (TC01) KHÔNG được xuất hiện trên PO mua hàng thông thường
  - Journal ALL BU (CCDC/VP/Tổng hợp/Khác): ràng buộc nằm ở Tax và Deliver To

[2] Billing Journal × Deliver To:
  - Journal CCDC → Deliver To phải là kho CCDC
  - Journal XL (NTP/vật liệu DA/khấu trừ) → Deliver To phải là kho NVL
  - Journal VP → Deliver To phải là kho Văn phòng / dùng nội bộ
  - Kiểm tra từng order line nếu Deliver To khác nhau

[3] Billing Journal × Tax keyword (dùng ILIKE, KHÔNG kiểm tra %):
  Mapping JOURNAL → TAX BẮT BUỘC (chỉ áp dụng cho BU chuyên biệt):
  - HD hoạt động đầu tư             → Tax ILIKE '%GTGT Đầu tư%'
  - HD KDVT                         → Tax ILIKE '%GTGT KDVT%'
  - HD NTP / vật liệu DA / khấu trừ CĐT → Tax ILIKE '%GTGT Xây Lắp%'

  Mapping JOURNAL → TAX áp dụng cho TẤT CẢ BU (ALL BU — không báo lỗi cross-BU):
  - HD mua CCDC             → Tax ILIKE '%GTGT HHDV%'   → Deliver To: kho CCDC
  - HD mua đồ dùng VP       → Tax ILIKE '%GTGT HHDV%'   → Deliver To: Văn phòng/Nội bộ
  - HD mua hàng tổng hợp    → Tax ILIKE '%GTGT Khác%'
  - HD mua hàng khác        → Tax ILIKE '%GTGT Nhập khẩu%'

  QUAN TRỌNG — Không báo lỗi cross-BU với các loại tax sau:
  - GTGT TSCĐ: bất kỳ BU nào cũng có thể mua tài sản cố định → PASS
  - GTGT Khác: bất kỳ BU nào cũng có thể dùng → PASS
  - GTGT Nhập khẩu: bất kỳ BU nào cũng có thể nhập khẩu → PASS
  - GTGT HHDV: bất kỳ BU nào cũng có thể mua hàng hóa dịch vụ → PASS
  Ví dụ: XL01 mua xe ô tô (TSCĐ) dùng "HD mua CCDC" + "GTGT TSCĐ" → PASS, không phải lỗi.
  Lưu ý: Xe ô tô ≥30tr + dùng ≥1 năm = TSCĐ (TK 211), KHÔNG phải CCDC (TK 153)

[4] Analytic BU & Taxes:
  Kiểm tra ĐỒNG THỜI 2 điều kiện:
  a) Analytic BU khớp với BU của PO:
     - Mã analytic account phải thuộc đúng nhóm BU (XL* cho Xây lắp, DT* cho Đầu tư, KD* cho Kinh doanh...)
     - Không dùng analytic của BU khác
     - Tổng % phân bổ trên mỗi order line PHẢI = 100%
  b) Tax keyword khớp với Billing Journal theo bảng mapping:
     - Dùng ILIKE '%KEYWORD%', không kiểm tra % thực tế
     - Các loại ALL BU (GTGT TSCĐ, GTGT Khác, GTGT Nhập khẩu, GTGT HHDV) không báo lỗi cross-BU
  → PASS chỉ khi CẢ HAI điều kiện a) và b) đều đúng

Sau khi hoàn thành phân tích nội bộ ở bước 1 và 2, CHỈ trả về phần kết luận theo đúng format dưới đây.
KHÔNG trả về bước 1, bước 2, hay bất kỳ nội dung phân tích nào khác.
Bắt đầu câu trả lời NGAY bằng dòng "KẾT LUẬN:".

KẾT LUẬN: PO [mã PO] — [1 câu tóm tắt tổng thể: hợp lệ hoặc mô tả số lỗi và mức độ nghiêm trọng]

[Nếu có lỗi, liệt kê từng lỗi theo format:]
Lỗi 1 — [Tên loại lỗi] [Critical/Warning]: [Mô tả cụ thể: giá trị hiện tại là gì, phải là gì, theo ràng buộc nào]
Lỗi 2 — [Tên loại lỗi] [Critical/Warning]: [Mô tả cụ thể]
(bỏ qua phần này nếu không có lỗi)

Gợi ý xử lý:
1. [Hành động cụ thể cần làm ngay]
2. [Hành động cụ thể cần làm ngay]
3. [Kiểm tra mở rộng nếu cần]
4. Verify BU = Khớp với nghiệp vụ và Thuế

TỔNG KẾT: [HỢP LỆ / CẦN KIỂM TRA / LỖI]"""

        scenarios.append({
            "po_number": po_num,
            "source_file": header.get("_source_file", ""),
            "total_lines": len(lines),
            "prompt": prompt,
            "header_data": header,
            "analytic_summary": analytic_summary,
        })
        context.log.info(f"  [{header.get('_source_file','')}] PO={po_num}: {len(lines)} order lines")

    return scenarios


@dg.asset(group_name="po_validation")
def validate_with_gemini(context: dg.AssetExecutionContext, build_scenarios: List[Dict]) -> List[Dict]:
    """Goi Groq (Llama) de validate tung PO (1 lan / PO) — mien phi, 14400 req/ngay."""
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key or api_key.startswith("your-"):
        raise ValueError("Chua cau hinh GROQ_API_KEY trong file .env")

    client = Groq(api_key=api_key)
    GROQ_MODEL = "llama-3.3-70b-versatile"

    system_prompt = """Bạn là chuyên gia kiểm tra dữ liệu Purchase Order trong hệ thống Odoo ERP.
QUAN TRỌNG: Chỉ trả lời đúng format được yêu cầu. Không giải thích thêm, không dùng markdown, không thêm tiêu đề ngoài format.

Khi nhận thông tin một PO, thực hiện đúng 3 bước:

BƯỚC 1 — XÁC ĐỊNH NGHIỆP VỤ bằng bộ tín hiệu confidence:
  +3đ: Subcontractor Contract có dữ liệu → gần chắc chắn XL/ĐT
  +3đ: Analytic Distribution chứa mã BU (XL*, DT*, KD*...)
  +3đ: Tax name chứa keyword đúng nhóm
  +2đ: Deliver To chứa "NVL", "CCDC", "Hàng hóa"
  +2đ: Billing Journal thuộc nhóm đúng theo bảng ma trận
  +1đ: Business Type = Construction/Trading
  ≥9đ = Rất chắc chắn | 6–8đ = Khả năng cao | ≤5đ = Cần xác nhận

BƯỚC 2 — KIỂM TRA 4 RÀNG BUỘC:
  [1] Billing Journal × BU: Journal phải khớp với BU xác định từ bộ tín hiệu
  [2] Billing Journal × Deliver To: Journal CCDC→Kho CCDC; Journal XL→Kho NVL
  [3] Billing Journal × Tax keyword: dùng ILIKE '%KEYWORD%', KHÔNG kiểm tra %
  [4] Analytic BU & Taxes: Kiểm tra ĐỒNG THỜI
      a) Analytic Account phải thuộc đúng BU của PO (XL*→XL, DT*→DT, KD*→KD), tổng % = 100%
      b) Tax keyword phải khớp với Billing Journal theo bảng mapping
      → PASS chỉ khi CẢ HAI điều kiện đều đúng

BẢNG MAPPING JOURNAL → TAX KEYWORD:
  [BU chuyên biệt — chỉ đúng BU đó mới dùng]
  HD hoạt động đầu tư          → GTGT Đầu tư      (chỉ DT*)
  HD KDVT                      → GTGT KDVT         (chỉ KD*)
  HD NTP / vật liệu DA / khấu trừ CĐT → GTGT Xây Lắp (chỉ XL*)

  [ALL BU — bất kỳ BU nào cũng có thể dùng, KHÔNG báo lỗi cross-BU]
  HD mua CCDC                  → GTGT TSCĐ         (TẤT CẢ BU)
  HD mua đồ dùng VP            → GTGT HHDV         (TẤT CẢ BU)
  HD mua hàng tổng hợp         → GTGT Khác         (TẤT CẢ BU)
  HD mua hàng khác             → GTGT Nhập khẩu    (TẤT CẢ BU)

LƯU Ý ĐẶC BIỆT:
  - Xe ô tô ≥30tr + dùng ≥1 năm = TSCĐ (TK 211), tax phải là GTGT TSCĐ
  - "Nhật ký vay" (TC01) KHÔNG được xuất hiện trên PO mua hàng thông thường
  - Journal ALL BU (CCDC/VP/Tổng hợp/Khác): ràng buộc nằm ở Tax và Deliver To

BƯỚC 3 — Đánh giá từng chỉ tiêu: PASS / LỖI [Critical] / CẦN VERIFY
Trả lời bằng tiếng Việt, có cấu trúc rõ ràng."""

    results = []

    for scenario in build_scenarios:
        context.log.info(f"Validate PO={scenario['po_number']} ({scenario['total_lines']} lines)...")
        try:
            response = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": scenario["prompt"]},
                ],
                temperature=0.1,
                max_tokens=1024,
            )
            raw = response.choices[0].message.content

            # Chỉ lấy phần từ "KẾT LUẬN:" trở đi, bỏ phần phân tích bước 1-2
            if "KẾT LUẬN:" in raw:
                answer = raw[raw.index("KẾT LUẬN:"):].strip()
            else:
                answer = raw.strip()

            if "TỔNG KẾT: LỖI" in answer or "LỖI [CRITICAL]" in answer:
                status = "LỖI"
            elif "TỔNG KẾT: HỢP LỆ" in answer:
                status = "HỢP LỆ"
            else:
                status = "CẦN KIỂM TRA"

            context.log.info(f"  -> {status}")
            context.log.info(f"  -> {answer[:300]}...")

            # 1 dong duy nhat cho 1 PO
            h = scenario["header_data"]
            results.append({
                "Source File":            scenario["source_file"],
                "Ma PO":                  scenario["po_number"],
                "Vendor":                 h.get("Vendor", ""),
                "Buyer":                  h.get("Buyer", ""),
                "Buyer/Department":       h.get("Buyer/Department", ""),
                "Business Type":          h.get("Business Type", ""),
                "Billing Journal":        h.get("Billing Journal", ""),
                "Subcontractor Contract": h.get("Subcontractor Contract Project", ""),
                "Deliver To":             h.get("Deliver To", ""),
                "Taxes":                  h.get("Taxes", ""),
                "Analytic Accounts":      scenario["analytic_summary"],
                "So Order Lines":         scenario["total_lines"],
                "Trang thai":             status,
                "Ket qua Gemini":         answer,
            })

        except Exception as e:
            context.log.warning(f"  -> Loi Gemini API: {e}")
            h = scenario["header_data"]
            results.append({
                "Source File":            scenario["source_file"],
                "Ma PO":                  scenario["po_number"],
                "Vendor":                 h.get("Vendor", ""),
                "Buyer":                  h.get("Buyer", ""),
                "Buyer/Department":       h.get("Buyer/Department", ""),
                "Business Type":          h.get("Business Type", ""),
                "Billing Journal":        h.get("Billing Journal", ""),
                "Subcontractor Contract": h.get("Subcontractor Contract Project", ""),
                "Deliver To":             h.get("Deliver To", ""),
                "Taxes":                  h.get("Taxes", ""),
                "Analytic Accounts":      scenario["analytic_summary"],
                "So Order Lines":         scenario["total_lines"],
                "Trang thai":             "LOI_API",
                "Ket qua Gemini":         str(e),
            })

    hop_le = sum(1 for r in results if r["Trang thai"] == "HỢP LỆ")
    can_kt = sum(1 for r in results if r["Trang thai"] == "CẦN KIỂM TRA")
    loi    = sum(1 for r in results if "LỖI" in r["Trang thai"])
    context.log.info(f"Tong ket: {hop_le} hop le | {can_kt} can kiem tra | {loi} loi")
    return results


@dg.asset(group_name="po_validation")
def save_validation_results(context: dg.AssetExecutionContext, validate_with_gemini: List[Dict]) -> str:
    """Luu ket qua validation ra Excel co mau sac."""
    from openpyxl.styles import PatternFill, Font, Alignment

    df = pd.DataFrame(validate_with_gemini)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(VALIDATION_DIR, f"po_validation_result_{timestamp}.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Validation")
        ws = writer.sheets["Validation"]

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F9D58")  # xanh Gemini
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        status_col = df.columns.get_loc("Trang thai") + 1
        colors = {"HỢP LỆ": "C6EFCE", "CẦN KIỂM TRA": "FFEB9C", "LỖI": "FFC7CE"}
        for row_idx in range(2, ws.max_row + 1):
            status_val = str(ws.cell(row=row_idx, column=status_col).value or "").upper()
            fill_color = next((c for k, c in colors.items() if k in status_val), None)
            if fill_color:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill("solid", fgColor=fill_color)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col), 60
            ) + 2

    context.log.info(f"Da luu: {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════
# JOB 3: VENDOR BILL VALIDATION (Python Rules Engine)
# ══════════════════════════════════════════════════════════

import re
import json as _json
from datetime import date as _date

VENDOR_BILL_DIR = os.path.join(os.path.dirname(__file__), "validation_data", "vendor_bill")

# Column name aliases (try each name in order until found in DataFrame)
_VB_COL_ALIASES = {
    "journal_name":          ["journal_name", "Journal", "Nhật ký", "So Nhat Ky"],
    "activity_name":         ["activity_name", "Activity", "Hoạt động", "Hoat Dong"],
    "invoice_serial":        ["invoice_serial", "Invoice Serial", "Invoice/Bill Serial Number",
                              "Ký hiệu HĐ", "Ký hiệu hóa đơn", "Ký hiệu", "Serial Number", "Serial"],
    "invoice_number":        ["invoice_number", "Invoice Number", "Số HĐ", "So HD", "Bill Number", "Reference", "Ref"],
    "invoice_date":          ["invoice_date", "Invoice Date", "Invoice/Bill Date", "Ngày HĐ", "Ngay HD", "Bill Date"],
    "payment_reference":     ["payment_reference", "Payment Reference", "Diễn giải TT"],
    "attachment_count":      ["attachment_count", "Attachment Count", "Số đính kèm", "Attachments"],
    "name":                  ["name", "Name", "Diễn giải", "Label", "Description"],
    "account_code":          ["account_code", "Account Code", "Tài khoản", "Account"],
    "analytic_account_name": ["analytic_account_name", "Analytic Account", "Khoản mục", "Analytic"],
    "tax_description":       ["tax_description", "Tax Description", "Diễn giải thuế"],
    "tax_name":              ["tax_name", "Invoice lines/Taxes", "Tax Name", "Thuế", "Tax", "Taxes"],
}

# Header-level columns to forward-fill (merged cells in Excel)
VB_HEADER_COLS = [
    "journal_name", "activity_name", "invoice_serial", "invoice_number",
    "invoice_date", "payment_reference", "attachment_count", "tax_description",
]


def _resolve_col(df_cols: list, field: str) -> str | None:
    """Trả về tên cột thực tế trong DataFrame tương ứng với field chuẩn."""
    for alias in _VB_COL_ALIASES.get(field, [field]):
        if alias in df_cols:
            return alias
    return None


def _get(row: Dict, field: str) -> str:
    """Lấy giá trị chuẩn hoá từ row theo field name (đã được map sẵn vào key chuẩn)."""
    return str(row.get(field) or "").strip()


def _UNUSED_validate_one_bill(bill_key: str, header: Dict, lines: List[Dict]) -> Dict:
    """DEPRECATED — replaced by Groq AI validation."""
    today   = _date.today().isoformat()
    errors  = []   # list of (rule_id, label, detail)
    r_pass  = {}   # rule_id → label khi PASS
    r_na    = {}   # rule_id → label khi N/A (rule không áp dụng)

    def err(rule_id: str, label: str, detail: str):
        errors.append((rule_id, label, detail))

    def ok(rule_id: str, label: str):
        r_pass[rule_id] = label

    def na(rule_id: str, label: str):
        r_na[rule_id] = label

    # ══ GROUP A — HEADER & LEGAL ═══════════════════════════

    # Rule 1 — Attachment Count
    try:
        att = int(float(_get(header, "attachment_count") or 0))
    except Exception:
        att = 0
    if att <= 0:
        err("R1", "Attachment Count",
            f"attachment_count = {att} (phải > 0). Vui lòng đính kèm file chứng từ hóa đơn.")
    else:
        ok("R1", f"Attachment Count — {att} file đính kèm")

    # Rule 2 — Invoice Serial (TT78): [1-6][CK][0-9]{2}[A-Z]{3}, ví dụ: 1C26TBP
    serial = _get(header, "invoice_serial")
    if not serial:
        err("R2", "Serial Number", "Trống — bắt buộc phải có ký hiệu hóa đơn (ví dụ: 1C26TBP).")
    elif not re.match(r'^[1-6][CK][0-9]{2}[A-Z]{3}$', serial):
        err("R2", "Serial Number",
            f"'{serial}' không đúng định dạng TT78 (^[1-6][CK][0-9]{{2}}[A-Z]{{3}}$). Ví dụ: '1C26TBP'.")
    else:
        ok("R2", f"Serial Number — {serial}")

    # Rule 3 — Invoice Number: số nguyên, dấu /, hoặc "/"
    inv_num = _get(header, "invoice_number")
    if not inv_num:
        err("R3", "Bill Number", "Trống — bắt buộc phải có số hóa đơn.")
    elif not re.match(r'^[0-9/]{1,20}$', inv_num):
        err("R3", "Bill Number",
            f"'{inv_num}' chỉ được chứa chữ số và dấu '/' (ví dụ: '0000001' hoặc '001/2024').")
    else:
        ok("R3", f"Bill Number — {inv_num}")

    # Rule 4 — Invoice Date
    inv_date = _get(header, "invoice_date")
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            from datetime import datetime as _dt
            inv_date = _dt.strptime(inv_date, fmt).strftime("%Y-%m-%d")
            break
        except Exception:
            pass
    if not re.match(r'^\d{4}-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01])$', inv_date):
        err("R4", "Bill Date",
            f"'{inv_date}' không đúng định dạng (YYYY-MM-DD hoặc DD/MM/YYYY).")
    elif inv_date > today:
        err("R4", "Bill Date",
            f"'{inv_date}' là ngày tương lai (hôm nay: {today}). Hóa đơn không được ghi ngày tương lai.")
    else:
        ok("R4", f"Bill Date — {inv_date}")

    # Rule 5 — Payment Reference
    pay_ref = _get(header, "payment_reference")
    if not pay_ref:
        err("R5", "Payment Reference", "Trống — bắt buộc phải có diễn giải thanh toán.")
    else:
        ok("R5", f"Payment Reference — có giá trị")

    # ══ GROUP B — LINE ITEMS ════════════════════════════════

    label_errors, tax_dir_errors = [], []
    for i, line in enumerate(lines, 1):
        name_val = _get(line, "name")
        if not name_val:
            label_errors.append(f"Dòng {i}: Diễn giải trống")

        tax_name = _get(line, "tax_name")
        if tax_name and not re.match(r'^(?=.*\bIN\b)(?!.*\bOUT\b).*$', tax_name):
            tax_dir_errors.append(f"Dòng {i}: '{tax_name}' — phải chứa IN, không được có OUT")

    if label_errors:
        err("R6", "Label (Diễn giải)", "; ".join(label_errors))
    else:
        ok("R6", "Label — tất cả dòng có diễn giải")

    if tax_dir_errors:
        err("R7", "Tax Direction (IN/OUT)", "; ".join(tax_dir_errors))
    else:
        ok("R7", "Tax Direction — tất cả dòng dùng thuế đầu vào (IN)")

    # ══ GROUP C — CROSS-VALIDATION ══════════════════════════

    journal  = _get(header, "journal_name")
    activity = _get(header, "activity_name")

    bds_errors, xl_errors, kdvt_errors = [], [], []
    for i, line in enumerate(lines, 1):
        analytic = _get(line, "analytic_account_name")
        tax_name = _get(line, "tax_name")

        # Rule 8 — Đầu tư BĐS
        if re.search(r'(?i)(Đầu tư:\s*(Xây lắp|Bất Động Sản|Cho thuê Văn phòng|Năng lượng)|BĐS)', analytic):
            if not re.search(r'(?i)GTGT\s*Đầu tư', tax_name):
                bds_errors.append(f"Dòng {i}: analytic='{analytic}' nhưng tax='{tax_name}' thiếu 'GTGT Đầu tư'")

        # Rule 9 — Xây lắp
        if re.search(r'(?i)Xây lắp:\s*(Dân dụng|Công nghiệp|Giao thông\s*&\s*Hạ tầng)', analytic):
            if not re.search(r'(?i)GTGT\s*Xây Lắp', tax_name):
                xl_errors.append(f"Dòng {i}: analytic='{analytic}' nhưng tax='{tax_name}' thiếu 'GTGT Xây Lắp'")

        # Rule 10 — KDVT
        if re.search(r'(?i)Kinh doanh:\s*Vật tư xây dựng', analytic):
            if not re.search(r'(?i)Hoá đơn mua hàng hóa KDVT', journal):
                kdvt_errors.append(f"Dòng {i}: journal='{journal}' phải là 'Hoá đơn mua hàng hóa KDVT'")
            elif not re.search(r'(?i)GTGT\s*KDVT', tax_name):
                kdvt_errors.append(f"Dòng {i}: tax='{tax_name}' thiếu 'GTGT KDVT'")

    if bds_errors:
        err("R8", "Đầu tư BĐS × Tax", "; ".join(bds_errors))
    else:
        na("R8", "Đầu tư BĐS × Tax — không có analytic BĐS") if not any(
            re.search(r'(?i)(Đầu tư.*BĐS|BĐS)', _get(ln, "analytic_account_name")) for ln in lines
        ) else ok("R8", "Đầu tư BĐS × Tax — PASS")

    if xl_errors:
        err("R9", "Xây lắp × Tax", "; ".join(xl_errors))
    else:
        na("R9", "Xây lắp × Tax — không có analytic Xây lắp") if not any(
            re.search(r'(?i)Xây lắp:', _get(ln, "analytic_account_name")) for ln in lines
        ) else ok("R9", "Xây lắp × Tax — PASS")

    if kdvt_errors:
        err("R10", "KDVT × Journal × Tax", "; ".join(kdvt_errors))
    else:
        na("R10", "KDVT × Journal × Tax — không có analytic KDVT") if not any(
            re.search(r'(?i)Kinh doanh:', _get(ln, "analytic_account_name")) for ln in lines
        ) else ok("R10", "KDVT × Journal × Tax — PASS")

    # Rule 11 & 12 — Hoạt động Đầu tư
    is_dt = bool(re.search(r'(?i)HOẠT ĐỘNG ĐẦU TƯ', activity)) or any(
        re.search(r'(?i)HOẠT ĐỘNG ĐẦU TƯ', _get(ln, "analytic_account_name")) for ln in lines
    )
    if is_dt:
        allowed = (r'(?i)^(Hóa đơn mua hàng cho hoạt động đầu tư|Hóa đơn mua hàng - KDDV'
                   r'|Hóa đơn mua CCDC|Hóa đơn mua đồ dùng văn phòng'
                   r'|Hóa đơn mua hàng tổng hợp|Hóa đơn mua hàng khác)$')
        if not re.match(allowed, journal):
            err("R11", "Hoạt động Đầu tư × Journal",
                f"journal='{journal}' không thuộc danh sách cho phép. "
                f"Dùng: Hóa đơn mua hàng cho hoạt động đầu tư / CCDC / VP / Tổng hợp / Khác.")
        else:
            ok("R11", f"Hoạt động Đầu tư × Journal — {journal}")

        sub = (r'(?i)^(Hóa đơn mua CCDC|Hóa đơn mua đồ dùng văn phòng'
               r'|Hóa đơn mua hàng tổng hợp|Hóa đơn mua hàng khác)$')
        if re.match(sub, journal):
            sub_errs = []
            for i, line in enumerate(lines, 1):
                tn = _get(line, "tax_name")
                if not re.search(r'(?i)GTGT\s+(HHDV|Khác|Nhập khẩu|TSCĐ)', tn):
                    sub_errs.append(f"Dòng {i}: '{tn}' không khớp GTGT(HHDV|Khác|Nhập khẩu|TSCĐ)")
            if sub_errs:
                err("R12", "Nhật ký ngách × Tax", "; ".join(sub_errs))
            else:
                ok("R12", "Nhật ký ngách × Tax — PASS")
        else:
            na("R12", "Nhật ký ngách × Tax — không áp dụng cho journal này")
    else:
        na("R11", "Hoạt động Đầu tư × Journal — không có activity Đầu tư")
        na("R12", "Nhật ký ngách × Tax — không có activity Đầu tư")

    # ══ BUILD CONCLUSION ════════════════════════════════════

    def fmt_rule(rid, label, tag):
        pad = " " * max(0, 32 - len(label))
        return f"  {rid:<4}: {label}{pad}→ {tag}"

    lines_out = []

    # Group A
    lines_out.append("GROUP A — Pháp lý & Định danh:")
    for rid in ["R1","R2","R3","R4","R5"]:
        if rid in r_pass:
            lines_out.append(fmt_rule(rid, r_pass[rid].split(" — ")[0], f"PASS — {r_pass[rid].split(' — ',1)[-1]}"))
        else:
            e = next((e for e in errors if e[0]==rid), None)
            if e: lines_out.append(fmt_rule(rid, e[1], f"LỖI — {e[2]}"))

    # Group B
    lines_out.append("\nGROUP B — Chi tiết dòng:")
    for rid in ["R6","R7"]:
        if rid in r_pass:
            lines_out.append(fmt_rule(rid, r_pass[rid].split(" — ")[0], f"PASS — {r_pass[rid].split(' — ',1)[-1]}"))
        else:
            e = next((e for e in errors if e[0]==rid), None)
            if e: lines_out.append(fmt_rule(rid, e[1], f"LỖI — {e[2]}"))

    # Group C
    lines_out.append("\nGROUP C — Xác thực chéo:")
    for rid in ["R8","R9","R10","R11","R12"]:
        if rid in r_pass:
            lines_out.append(fmt_rule(rid, r_pass[rid].split(" — ")[0], f"PASS — {r_pass[rid].split(' — ',1)[-1]}"))
        elif rid in r_na:
            lines_out.append(fmt_rule(rid, r_na[rid].split(" — ")[0], f"N/A  — {r_na[rid].split(' — ',1)[-1]}"))
        else:
            e = next((e for e in errors if e[0]==rid), None)
            if e: lines_out.append(fmt_rule(rid, e[1], f"LỖI — {e[2]}"))

    n_err = len(errors)
    if n_err == 0:
        summary = "HỢP LỆ — tất cả 12 quy tắc đều thỏa mãn."
        status  = "VALID"
    else:
        failed_labels = ", ".join(e[1] for e in errors)
        summary = f"INVALID — {n_err} lỗi: {failed_labels}"
        status  = "INVALID"

    conclusion = (
        f"KẾT LUẬN: {summary}\n\n"
        + "\n".join(lines_out)
        + f"\n\nTỔNG KẾT: {'HỢP LỆ' if status == 'VALID' else 'INVALID — cần xử lý trước khi hạch toán'}"
    )

    return {
        "status":      status,
        "error_type":  errors[0][1] if errors else "N/A",
        "failed_rule": ", ".join(e[0] for e in errors) if errors else None,
        "message":     conclusion,
    }


# ── Assets ────────────────────────────────────────────────

@dg.asset(group_name="vendor_bill")
def load_vendor_bill_data(context: dg.AssetExecutionContext) -> List[Dict]:
    """Loop tất cả file Excel trong vendor_bill_data/, map cột chuẩn, forward-fill header."""
    os.makedirs(VENDOR_BILL_DIR, exist_ok=True)
    xlsx_files = [
        f for f in glob.glob(os.path.join(VENDOR_BILL_DIR, "*.xlsx"))
        if "result" not in os.path.basename(f).lower()
    ]
    if not xlsx_files:
        raise FileNotFoundError(f"Không tìm thấy file Excel nào trong {VENDOR_BILL_DIR}")

    all_records: List[Dict] = []

    for path in xlsx_files:
        filename = os.path.basename(path)
        context.log.info(f"Đọc: {filename}")
        df = pd.read_excel(path, sheet_name=0, dtype=str)
        df_cols = list(df.columns)

        # Map tên cột thực → key chuẩn
        col_map = {}
        for field in list(_VB_COL_ALIASES.keys()):
            real = _resolve_col(df_cols, field)
            if real:
                col_map[real] = field

        df = df.rename(columns=col_map)
        mapped   = list(col_map.values())
        unmapped = [c for c in df_cols if c not in col_map]
        context.log.info(f"  Cột nhận dạng : {mapped}")
        if unmapped:
            context.log.info(f"  Cột chưa map  : {unmapped}")

        # Cột A (cột đầu tiên) có giá trị → đó là dòng HEADER của hóa đơn mới
        # Các dòng bên dưới (cột A trống) → là BILL LINES, data từ cột L trở đi
        first_col = df_cols[0]
        df["_is_header"] = df[first_col].notna() & (df[first_col].astype(str).str.strip() != "")

        # Forward-fill header columns xuống các dòng bill lines
        ffill_cols = [c for c in VB_HEADER_COLS if c in df.columns]
        df[ffill_cols] = df[ffill_cols].ffill()

        # Chỉ giữ lại dòng có data ở line columns (name, tax_name, analytic...)
        line_cols = [c for c in ["name", "account_code", "analytic_account_name", "tax_name"] if c in df.columns]
        if line_cols:
            has_line_data = df[line_cols].notna().any(axis=1)
            df_lines = df[has_line_data].copy()
        else:
            df_lines = df.copy()

        context.log.info(f"  → {len(df)} dòng gốc | {df['_is_header'].sum()} header | {len(df_lines)} bill lines")

        df_lines["_source_file"] = filename
        records = df_lines.where(pd.notna(df_lines), None).to_dict(orient="records")
        all_records.extend(records)

        bills = df_lines["invoice_number"].dropna().unique() if "invoice_number" in df_lines.columns else []
        context.log.info(f"  → {len(bills)} hóa đơn: {list(bills)}")

    context.log.info(f"Tổng: {len(all_records)} dòng từ {len(xlsx_files)} file")
    return all_records


# 12 cột tiêu chí — mỗi cột = 1 rule
_VB_CRITERIA_BASE = [
    "R1_Attachment",
    "R2_Serial_HĐ",
    "R3_Số_HĐ",
    "R4_Ngày_HĐ",
    "R5_Diễn_giải_TT",
    "R6_Diễn_giải_thuế",
    "R7_Label_dòng",
    "R8_Chiều_thuế",
    "R9_ĐT_BĐS×Tax",
    "R10_Xây_lắp×Tax",
    "R11_KDVT×Journal",
    "R12_HĐ_ĐT×Journal",
]

# Mỗi R tách thành 2 cột: verdict (emoji) + chi tiết
_VB_CRITERIA_COLS = []
for _b in _VB_CRITERIA_BASE:
    _VB_CRITERIA_COLS.append(_b)
    _VB_CRITERIA_COLS.append(_b + "_CT")

# Regex patterns để nhận dạng từng dòng trong kết quả Groq
_VB_CRITERIA_PATTERNS = [
    (0,  r'1\.\s*Attachment'),
    (1,  r'2\.\s*(Ký hiệu|Serial)'),
    (2,  r'3\.\s*(Số hóa đơn|Bill Number|Số HĐ)'),
    (3,  r'4\.\s*(Ngày|Bill Date)'),
    (4,  r'5\.\s*(Diễn giải thanh toán|Payment)'),
    (5,  r'6\.\s*(Diễn giải thuế|Tax Desc)'),
    (6,  r'7\.\s*(Diễn giải hàng|Label)'),
    (7,  r'8\.\s*(Chiều thuế|IN/OUT|Tax Direction)'),
    (8,  r'9\.\s*(Đầu tư BĐS|BĐS)'),
    (9,  r'10\.\s*(Xây lắp)'),
    (10, r'11\.\s*(KDVT)'),
    (11, r'12\.\s*(Hoạt động ĐT|HĐ ĐT|Đầu tư.*Journal|Hoạt động Đầu tư)'),
]


def _extract_verdict_split(line: str):
    """Trích xuất (verdict_emoji, detail) từ 1 dòng kết quả Groq.
    Trả về tuple: (str verdict, str detail)
    """
    line = line.strip()
    # Xác định verdict tag
    if "PASS" in line.upper():
        tag = "✅ PASS"
    elif "LỖI" in line.upper() or "INVALID" in line.upper() or "❌" in line:
        tag = "❌ LỖI"
    elif "N/A" in line.upper() or "⚪" in line:
        tag = "⚪ N/A"
    else:
        tag = line[:60]

    # Lấy phần diễn giải sau dấu — hoặc :
    # Dòng có dạng:  "  1. Attachment Count : ❌ LỖI — Giá trị ..."
    # Tìm vị trí sau verdict tag trong line, rồi lấy phần còn lại
    detail = ""
    # Thử tách theo " — " sau verdict
    for sep in [" — ", " —", "—"]:
        idx = line.find(sep)
        if idx != -1:
            candidate = line[idx + len(sep):].strip()
            # Bỏ prefix PASS/LỖI/N/A lặp nếu còn
            for prefix in ["✅ PASS", "❌ LỖI", "⚪ N/A", "PASS", "LỖI", "N/A"]:
                if candidate.upper().startswith(prefix.upper()):
                    candidate = candidate[len(prefix):].lstrip(" —:")
            detail = candidate.strip()
            if detail:
                break

    # Fallback: tách theo ": " cuối cùng trong dòng
    if not detail:
        for sep in [": "]:
            parts = line.rsplit(sep, 1)
            if len(parts) == 2:
                candidate = parts[1].strip()
                for prefix in ["✅ PASS", "❌ LỖI", "⚪ N/A", "PASS", "LỖI", "N/A"]:
                    if candidate.upper().startswith(prefix.upper()):
                        candidate = candidate[len(prefix):].lstrip(" —:")
                detail = candidate.strip()

    return tag, detail


def _parse_vb_criteria(message: str) -> Dict:
    """Parse kết quả Groq thành dict 24 cột tiêu chí (12 verdict + 12 chi tiết)."""
    result = {col: "" for col in _VB_CRITERIA_COLS}
    lines  = message.splitlines()

    for line in lines:
        line_clean = line.strip()
        if not line_clean:
            continue
        for col_idx, pattern in _VB_CRITERIA_PATTERNS:
            if re.search(pattern, line_clean, re.IGNORECASE):
                verdict, detail = _extract_verdict_split(line_clean)
                base_col = _VB_CRITERIA_BASE[col_idx]
                result[base_col]          = verdict
                result[base_col + "_CT"]  = detail
                break

    return result


@dg.asset(group_name="vendor_bill")
def build_vb_scenarios(context: dg.AssetExecutionContext, load_vendor_bill_data: List[Dict]) -> List[Dict]:
    """Gom lines theo hóa đơn, build JSON payload + prompt cho Groq — 1 scenario / bill."""
    bill_groups: Dict[str, List[Dict]] = {}
    for row in load_vendor_bill_data:
        inv_num = str(row.get("invoice_number") or "").strip()
        src     = str(row.get("_source_file") or "").strip()
        key     = src if inv_num in ("", "/") else f"{src}::{inv_num}"
        bill_groups.setdefault(key, []).append(row)

    context.log.info(f"Tìm thấy {len(bill_groups)} hóa đơn")
    scenarios = []

    for bill_key, lines in bill_groups.items():
        header  = lines[0]
        inv_num = str(header.get("invoice_number") or "?")
        src     = str(header.get("_source_file") or "")

        bill_json = {
            "header": {
                "journal_name":      _get(header, "journal_name"),
                "activity_name":     _get(header, "activity_name"),
                "invoice_serial":    _get(header, "invoice_serial"),
                "invoice_number":    inv_num,
                "invoice_date":      _get(header, "invoice_date"),
                "payment_reference": _get(header, "payment_reference"),
                "attachment_count":  _get(header, "attachment_count") or "0",
                "tax_description":   _get(header, "tax_description"),
            },
            "lines": [
                {
                    "line_id":               i,
                    "name":                  _get(line, "name"),
                    "account_code":          _get(line, "account_code"),
                    "analytic_account_name": _get(line, "analytic_account_name"),
                    "tax_name":              _get(line, "tax_name"),
                }
                for i, line in enumerate(lines, 1)
            ],
        }

        scenarios.append({
            "bill_key":       bill_key,
            "source_file":    src,
            "invoice_number": inv_num,
            "total_lines":    len(lines),
            "header_data":    header,
            "prompt":         _json.dumps(bill_json, ensure_ascii=False, indent=2),
        })
        context.log.info(f"  [{src}] Bill #{inv_num}: {len(lines)} dòng")

    return scenarios


@dg.asset(group_name="vendor_bill")
def validate_vendor_bills(context: dg.AssetExecutionContext, build_vb_scenarios: List[Dict]) -> List[Dict]:
    """RAG + Groq: query ChromaDB lấy context từ index_documents_job, gửi Groq validate từng bill."""
    groq_key   = os.getenv("GROQ_API_KEY")
    gemini_key = os.getenv("GEMINI_API_KEY")
    if not groq_key or groq_key.startswith("your-"):
        raise ValueError("Chưa cấu hình GROQ_API_KEY trong file .env")
    if not gemini_key or gemini_key.startswith("your-"):
        raise ValueError("Chưa cấu hình GEMINI_API_KEY trong file .env")

    groq_client   = Groq(api_key=groq_key)
    gemini_client = google_genai.Client(api_key=gemini_key)
    GROQ_MODEL    = "llama-3.3-70b-versatile"

    # Kết nối ChromaDB đã được index bởi index_documents_job
    chroma     = chromadb.PersistentClient(path=CHROMA_DIR)
    collection = chroma.get_collection(name=COLLECTION_NAME)
    context.log.info(f"ChromaDB: {collection.count()} chunks sẵn sàng")

    system_prompt = """Bạn là Chuyên gia Kế toán Odoo ERP và Kế toán trưởng chuyên về chuẩn Kế toán Việt Nam (VAS & Thông tư 78).

Nhiệm vụ: nhận dữ liệu JSON của một Vendor Bill cùng tài liệu nghiệp vụ liên quan, kiểm tra TUẦN TỰ đủ 12 chỉ số bên dưới, trả về kết luận bằng tiếng Việt.

Sau khi phân tích, CHỈ trả về phần kết luận theo đúng format. KHÔNG giải thích thêm, KHÔNG markdown, bắt đầu NGAY bằng "KẾT LUẬN:".

=== FORMAT KẾT QUẢ ===

Nếu HỢP LỆ:
KẾT LUẬN: ✅ Hóa đơn #[số HĐ] hợp lệ — không cần điều chỉnh.

GROUP A — Pháp lý & Định danh:
  1. Attachment Count      : ✅ PASS — [chi tiết]
  2. Ký hiệu HĐ (Serial)  : ✅ PASS — [giá trị]
  3. Số hóa đơn           : ✅ PASS — [giá trị]
  4. Ngày hóa đơn         : ✅ PASS — [giá trị]
  5. Diễn giải thanh toán : ✅ PASS
  6. Diễn giải thuế       : ✅ PASS

GROUP B — Chi tiết dòng:
  7. Diễn giải hàng HĐ   : ✅ PASS — tất cả dòng có diễn giải
  8. Chiều thuế (IN/OUT)  : ✅ PASS — tất cả dòng dùng thuế đầu vào (IN)

GROUP C — Xác thực chéo:
  9.  Đầu tư BĐS × Tax   : ✅ PASS — [chi tiết] | ⚪ N/A — [lý do không áp dụng]
  10. Xây lắp × Tax       : ✅ PASS — [chi tiết] | ⚪ N/A — [lý do]
  11. KDVT × Journal × Tax: ✅ PASS — [chi tiết] | ⚪ N/A — [lý do]
  12. Hoạt động ĐT × Journal: ✅ PASS — [chi tiết] | ⚪ N/A — [lý do]

TỔNG KẾT: HỢP LỆ

---

Nếu CÓ LỖI:
KẾT LUẬN: ❌ Hóa đơn #[số HĐ] không hợp lệ (phát hiện [N] lỗi):

GROUP A — Pháp lý & Định danh:
  1. Attachment Count      : ✅ PASS — [chi tiết] | ❌ LỖI — [mô tả lỗi cụ thể]
  2. Ký hiệu HĐ (Serial)  : ✅ PASS — [giá trị] | ❌ LỖI — [mô tả lỗi]
  3. Số hóa đơn           : ✅ PASS | ❌ LỖI — [mô tả lỗi]
  4. Ngày hóa đơn         : ✅ PASS | ❌ LỖI — [mô tả lỗi]
  5. Diễn giải thanh toán : ✅ PASS | ❌ LỖI — [mô tả lỗi]
  6. Diễn giải thuế       : ✅ PASS | ❌ LỖI — [mô tả lỗi]

GROUP B — Chi tiết dòng:
  7. Diễn giải hàng HĐ   : ✅ PASS | ❌ LỖI — Dòng [N]: [mô tả]
  8. Chiều thuế (IN/OUT)  : ✅ PASS | ❌ LỖI — Dòng [N]: '[giá trị]' chứa OUT

GROUP C — Xác thực chéo:
  9.  Đầu tư BĐS × Tax   : ✅ PASS | ❌ LỖI — [mô tả] | ⚪ N/A — [lý do]
  10. Xây lắp × Tax       : ✅ PASS | ❌ LỖI — [mô tả] | ⚪ N/A — [lý do]
  11. KDVT × Journal × Tax: ✅ PASS | ❌ LỖI — [mô tả] | ⚪ N/A — [lý do]
  12. Hoạt động ĐT × Journal: ✅ PASS | ❌ LỖI — [mô tả] | ⚪ N/A — [lý do]

Hướng xử lý:
- [Hành động cụ thể 1]
- [Hành động cụ thể 2]

TỔNG KẾT: INVALID — [N] lỗi cần xử lý"""

    results = []

    for scenario in build_vb_scenarios:
        inv_num = scenario["invoice_number"]
        src     = scenario["source_file"]
        context.log.info(f"  Validate: [{src}] Bill #{inv_num} ({scenario['total_lines']} dòng)...")

        try:
            # ── Bước 1: Embed query để query ChromaDB
            query_text = f"Vendor Bill validation rules: {scenario['prompt'][:300]}"
            embed_result = gemini_client.models.embed_content(
                model="models/gemini-embedding-001",
                contents=query_text,
            )
            query_vector = embed_result.embeddings[0].values

            # ── Bước 2: Retrieve top-5 chunks liên quan từ ChromaDB
            rag_result = collection.query(
                query_embeddings=[query_vector],
                n_results=5,
                include=["documents", "metadatas"],
            )
            retrieved_docs = rag_result["documents"][0] if rag_result["documents"] else []
            rag_context = "\n\n---\n\n".join(retrieved_docs)
            context.log.info(f"    RAG: lấy {len(retrieved_docs)} chunks từ ChromaDB")

            # ── Bước 3: Build user prompt = context + bill data
            user_prompt = f"""=== TÀI LIỆU NGHIỆP VỤ (từ hệ thống training) ===
{rag_context}

=== DỮ LIỆU VENDOR BILL CẦN KIỂM TRA ===
{scenario['prompt']}"""

            # ── Bước 4: Gọi Groq
            response = groq_client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user",   "content": user_prompt},
                ],
                temperature=0.0,
                max_tokens=1024,
            )
            raw = response.choices[0].message.content.strip()

            if "KẾT LUẬN:" in raw:
                message = raw[raw.index("KẾT LUẬN:"):].strip()
            else:
                message = raw

            # Parse từng dòng tiêu chí → dict {R1: "PASS — ...", R2: "LỖI — ...", ...}
            criteria = _parse_vb_criteria(message)

            # Xác định status tổng từ TỔNG KẾT
            msg_upper = message.upper()
            if "TỔNG KẾT: HỢP LỆ" in msg_upper:
                status = "HỢP LỆ"
            elif "TỔNG KẾT: INVALID" in msg_upper:
                status = "INVALID"
            else:
                status = "CẦN KIỂM TRA"

            context.log.info(f"    → {status}")

        except Exception as e:
            context.log.warning(f"    → Lỗi API: {e}")
            message  = str(e)
            status   = "LỖI API"
            criteria = {k: "" for k in _VB_CRITERIA_COLS}

        h = scenario["header_data"]
        row = {
            "Source File":       src,
            "Invoice Number":    inv_num,
            "Invoice Serial":    str(h.get("invoice_serial") or ""),
            "Invoice Date":      str(h.get("invoice_date") or ""),
            "Journal":           str(h.get("journal_name") or ""),
            "Activity":          str(h.get("activity_name") or ""),
            "Payment Reference": str(h.get("payment_reference") or ""),
            "Attachment Count":  str(h.get("attachment_count") or ""),
            "So Lines":          scenario["total_lines"],
        }
        row.update(criteria)           # thêm 12 cột tiêu chí
        row["VB_Status"]    = status
        row["VB_Kết_luận"]  = message
        results.append(row)

    hop_le = sum(1 for r in results if r["VB_Status"] == "HỢP LỆ")
    invalid = sum(1 for r in results if r["VB_Status"] == "INVALID")
    can_kt  = sum(1 for r in results if r["VB_Status"] == "CẦN KIỂM TRA")
    context.log.info(f"Tổng kết: {hop_le} hợp lệ | {can_kt} cần kiểm tra | {invalid} invalid")
    return results


@dg.asset(group_name="vendor_bill")
def save_vb_results(context: dg.AssetExecutionContext, validate_vendor_bills: List[Dict]) -> str:
    """Giữ nguyên dữ liệu file gốc, bổ sung cột kết quả ở cuối."""
    from openpyxl.styles import PatternFill, Font, Alignment

    # Build lookup: (source_file, invoice_number) → result
    result_map: Dict[str, Dict] = {}
    for r in validate_vendor_bills:
        key = f"{r['Source File']}::{r['Invoice Number']}"
        result_map[key] = r

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(VENDOR_BILL_DIR, exist_ok=True)
    out_path = os.path.join(VENDOR_BILL_DIR, f"vb_validation_result_{timestamp}.xlsx")

    xlsx_files = [
        f for f in glob.glob(os.path.join(VENDOR_BILL_DIR, "*.xlsx"))
        if "result" not in os.path.basename(f).lower()
    ]

    RESULT_COLS = ["VB_Status", "VB_Failed_Rule", "VB_Kết_luận"]
    colors = {
        "HỢP LỆ":       "C6EFCE",
        "INVALID":      "FFC7CE",
        "CẦN KIỂM TRA": "FFEB9C",
        "LỖI API":      "D9D9D9",
    }

    all_df_results: List = []

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for path in xlsx_files:
            filename  = os.path.basename(path)
            sheet_name = filename[:28]   # Excel sheet name max 31 chars
            context.log.info(f"  Xử lý: {filename}")

            df_orig = pd.read_excel(path, sheet_name=0, dtype=str)

            # Tìm cột invoice_number trong file gốc để lookup kết quả
            inv_col = next(
                (c for c in df_orig.columns
                 if any(alias.lower() in c.lower()
                        for alias in ["Reference", "Invoice Number", "Ref", "invoice_number"])),
                None,
            )

            # Map kết quả cho từng dòng theo invoice_number
            def get_result_val(row, field):
                if inv_col is None:
                    return ""
                inv = str(row.get(inv_col) or "").strip()
                key = f"{filename}::{inv}" if inv not in ("", "/") else filename
                res = result_map.get(key, {})
                return res.get(field, "")

            # 12 cột tiêu chí
            for col in _VB_CRITERIA_COLS:
                df_orig[col] = df_orig.apply(lambda r, c=col: get_result_val(r, c), axis=1)
            # Cột tổng kết
            df_orig["VB_Status"]   = df_orig.apply(lambda r: get_result_val(r, "VB_Status"), axis=1)
            df_orig["VB_Kết_luận"] = df_orig.apply(lambda r: get_result_val(r, "VB_Kết_luận"), axis=1)

            df_orig.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            # Style header
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1A73E8")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            all_cols       = list(df_orig.columns)
            status_idx     = all_cols.index("VB_Status") + 1
            ket_luan_idx   = all_cols.index("VB_Kết_luận") + 1
            # Cột verdict (R1_Attachment, R2_... không có _CT) → tô màu, hẹp
            verdict_idxs   = [all_cols.index(c) + 1 for c in _VB_CRITERIA_BASE if c in all_cols]
            # Cột chi tiết (_CT) → rộng hơn, wrap
            detail_idxs    = [all_cols.index(c + "_CT") + 1 for c in _VB_CRITERIA_BASE if c + "_CT" in all_cols]

            # Tô màu cột verdict theo emoji
            cell_colors = {
                "✅ PASS": "C6EFCE",
                "❌ LỖI": "FFC7CE",
                "⚪ N/A":  "F2F2F2",
            }
            for row_idx in range(2, ws.max_row + 1):
                for col_i in verdict_idxs:
                    val = str(ws.cell(row=row_idx, column=col_i).value or "")
                    fc  = next((c for k, c in cell_colors.items() if k in val), None)
                    if fc:
                        ws.cell(row=row_idx, column=col_i).fill = PatternFill("solid", fgColor=fc)

                # Tô cột VB_Status
                status_val = str(ws.cell(row=row_idx, column=status_idx).value or "").upper()
                fc_status  = next((c for k, c in colors.items() if k.upper() in status_val), None)
                if fc_status:
                    ws.cell(row=row_idx, column=status_idx).fill = PatternFill("solid", fgColor=fc_status)

            # Auto-width
            for col in ws.columns:
                col_letter = col[0].column_letter
                col_i      = col[0].column
                if col_i == ket_luan_idx:
                    ws.column_dimensions[col_letter].width = 90
                    for cell in col[1:]:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif col_i in detail_idxs:
                    ws.column_dimensions[col_letter].width = 45
                    for cell in col[1:]:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif col_i in verdict_idxs:
                    ws.column_dimensions[col_letter].width = 12
                    for cell in col[1:]:
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    ws.column_dimensions[col_letter].width = min(
                        max(len(str(c.value or "")) for c in col), 30
                    ) + 2

            context.log.info(f"  → Sheet '{sheet_name}': {len(df_orig)} dòng + 26 cột kết quả")

            # Tích lũy dữ liệu để build summary
            all_df_results.append(df_orig)

        # ── SUMMARY SHEET (bên trong with ExcelWriter) ────────
        if all_df_results:
            from openpyxl.styles import PatternFill as _PF, Font as _Font, Alignment as _Align

            df_all = pd.concat(all_df_results, ignore_index=True)

            # Tìm cột "Created By"
            created_by_col = next(
                (c for c in df_all.columns
                 if any(alias.lower() in c.lower()
                        for alias in ["created by", "created_by", "người tạo", "nguoi tao", "Responsible"])),
                None,
            )

            rule_labels = {
                "R1_Attachment":     "R1 — Attachment Count (Chứng từ đính kèm)",
                "R2_Serial_HĐ":      "R2 — Ký hiệu hóa đơn (Serial TT78)",
                "R3_Số_HĐ":          "R3 — Số hóa đơn",
                "R4_Ngày_HĐ":        "R4 — Ngày hóa đơn",
                "R5_Diễn_giải_TT":   "R5 — Diễn giải thanh toán",
                "R6_Diễn_giải_thuế": "R6 — Diễn giải thuế",
                "R7_Label_dòng":     "R7 — Diễn giải dòng hàng",
                "R8_Chiều_thuế":     "R8 — Chiều thuế IN/OUT",
                "R9_ĐT_BĐS×Tax":    "R9 — Đầu tư BĐS × Tax",
                "R10_Xây_lắp×Tax":   "R10 — Xây lắp × Tax",
                "R11_KDVT×Journal":  "R11 — KDVT × Journal × Tax",
                "R12_HĐ_ĐT×Journal": "R12 — Hoạt động Đầu tư × Journal",
            }

            # ── Bảng 1: Số lỗi theo Rule ─────────────────────
            rule_rows = []
            for col, label in rule_labels.items():
                if col not in df_all.columns:
                    continue
                n_loi  = int(df_all[col].str.contains("❌", na=False).sum())
                n_pass = int(df_all[col].str.contains("✅", na=False).sum())
                n_na   = int(df_all[col].str.contains("⚪", na=False).sum())
                rule_rows.append({
                    "Quy tắc":      label,
                    "Số lỗi (❌)":  n_loi,
                    "Số PASS (✅)": n_pass,
                    "Số N/A (⚪)":  n_na,
                    "Tổng HĐ":     n_loi + n_pass + n_na,
                })
            df_rule = pd.DataFrame(rule_rows)

            # ── Bảng 2: Số lỗi theo Created By ───────────────
            if created_by_col:
                df_person = df_all.copy()
                df_person["_has_error"] = df_all["VB_Status"].str.upper().isin(["INVALID", "CẦN KIỂM TRA"])
                person_summary = (
                    df_person.groupby(created_by_col)
                    .agg(Tổng_HĐ=("VB_Status", "count"), HĐ_lỗi=("_has_error", "sum"))
                    .reset_index()
                    .rename(columns={
                        created_by_col: "Người tạo (Created By)",
                        "Tổng_HĐ": "Tổng HĐ",
                        "HĐ_lỗi":  "HĐ có lỗi",
                    })
                )
                person_summary["HĐ hợp lệ"] = person_summary["Tổng HĐ"] - person_summary["HĐ có lỗi"]
                person_summary["% lỗi"] = (
                    person_summary["HĐ có lỗi"] / person_summary["Tổng HĐ"] * 100
                ).round(1).astype(str) + "%"
                for col, label in rule_labels.items():
                    short = label.split("—")[0].strip()
                    if col in df_all.columns:
                        person_summary[short] = (
                            df_person.groupby(created_by_col)[col]
                            .apply(lambda s: int(s.str.contains("❌", na=False).sum()))
                            .values
                        )
            else:
                person_summary = pd.DataFrame({"Ghi chú": ["Không tìm thấy cột 'Created By' trong file"]})

            # ── Ghi sheet SUMMARY ─────────────────────────────
            df_rule.to_excel(writer, index=False, sheet_name="📊 SUMMARY")
            ws_sum = writer.sheets["📊 SUMMARY"]

            header_fill_blue  = _PF("solid", fgColor="1A73E8")
            header_fill_green = _PF("solid", fgColor="0F9D58")

            # Style header bảng 1
            for cell in ws_sum[1]:
                cell.font      = _Font(bold=True, color="FFFFFF")
                cell.fill      = header_fill_blue
                cell.alignment = _Align(horizontal="center", wrap_text=True)

            # Tô đỏ ô lỗi > 0
            for row_idx in range(2, len(df_rule) + 2):
                loi_val = ws_sum.cell(row=row_idx, column=2).value
                if loi_val and int(loi_val) > 0:
                    ws_sum.cell(row=row_idx, column=2).fill = _PF("solid", fgColor="FFC7CE")
                    ws_sum.cell(row=row_idx, column=2).font = _Font(bold=True, color="C00000")

            ws_sum.column_dimensions["A"].width = 50
            for cl in ["B", "C", "D", "E"]:
                ws_sum.column_dimensions[cl].width = 16

            # Bảng 2: người tạo
            start_row = len(df_rule) + 4
            title_cell = ws_sum.cell(row=start_row, column=1,
                                     value="THỐNG KÊ THEO NGƯỜI TẠO (Created By)")
            title_cell.font = _Font(bold=True, size=13, color="1A73E8")

            for c_idx, col_name in enumerate(person_summary.columns, 1):
                cell = ws_sum.cell(row=start_row + 1, column=c_idx, value=col_name)
                cell.font      = _Font(bold=True, color="FFFFFF")
                cell.fill      = header_fill_green
                cell.alignment = _Align(horizontal="center", wrap_text=True)

            for r_idx, data_row in enumerate(person_summary.itertuples(index=False),
                                             start=start_row + 2):
                for c_idx, val in enumerate(data_row, 1):
                    cell = ws_sum.cell(row=r_idx, column=c_idx, value=val)
                    if c_idx == 3 and val and int(val) > 0:
                        cell.fill = _PF("solid", fgColor="FFC7CE")
                        cell.font = _Font(bold=True, color="C00000")

            # Đưa sheet SUMMARY lên đầu
            wb = writer.book
            wb.move_sheet("📊 SUMMARY", offset=-(len(wb.sheetnames) - 1))

            context.log.info(f"  → Summary: {len(df_rule)} rules | {len(person_summary)} người tạo")

    context.log.info(f"Đã lưu: {out_path}")
    return out_path


# ══════════════════════════════════════════════════════════
# JOBS & DEFINITIONS
# ══════════════════════════════════════════════════════════

index_documents_job = dg.define_asset_job(
    name="index_documents_job",
    selection=dg.AssetSelection.groups("training"),
    description="Doc tai lieu trong data/, tao embeddings local, luu vao ChromaDB",
)

po_validation_job = dg.define_asset_job(
    name="po_validation_job",
    selection=dg.AssetSelection.groups("po_validation"),
    description="[Groq Llama] Validate PO Excel theo 4 rang buoc, luu ket qua",
)

vendor_bill_validation_job = dg.define_asset_job(
    name="vendor_bill_validation_job",
    selection=dg.AssetSelection.groups("vendor_bill"),
    description="[Python Rules Engine] Validate Vendor Bill Excel theo 12 quy tac TT78, luu ket qua",
)

defs = dg.Definitions(
    assets=[
        load_documents, chunk_documents, index_to_chromadb,
        load_po_data, build_scenarios, validate_with_gemini, save_validation_results,
        load_vendor_bill_data, build_vb_scenarios, validate_vendor_bills, save_vb_results,
    ],
    jobs=[index_documents_job, po_validation_job, vendor_bill_validation_job],
)
